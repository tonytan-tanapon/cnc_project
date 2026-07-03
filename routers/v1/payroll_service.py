from fastapi import APIRouter

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from fastapi.responses import StreamingResponse
router = APIRouter(
    prefix="/payroll_service",
    tags=["payroll_service"],
)

from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from collections import defaultdict
from datetime import datetime

from database import get_db
from models import (
    Employee,
    PayPeriod,
    TimeEntry,
    BreakEntry,
)
from models import TimeLeave
from sqlalchemy import func

router = APIRouter(
    prefix="/payroll_service",
    tags=["payroll_service"],
)


def week_key(dt):
    """
    Monday start week
    """
    weekday = dt.weekday()  # Monday=0
    monday = dt.date().fromordinal(
        dt.date().toordinal() - weekday
    )
    return str(monday)

def calculate_payroll_data(
    employee_id: int,
    pp_id: int,
    rate: float,
    ot_rate: float,
    show_payee: bool = False,
    db: Session = Depends(get_db),
):
    pp = (
        db.query(PayPeriod)
        .filter(PayPeriod.id == pp_id)
        .first()
    )

    if not pp:
        return {"error": "Pay period not found"}

    employee = (
        db.query(Employee)
        .filter(Employee.id == employee_id)
        .first()
    )

    if not employee:
        return {"error": "Employee not found"}

    employee_ids = [employee.id]

    if show_payee and employee.payroll_emp_id:

        payee = (
            db.query(Employee)
            .filter(
                Employee.id ==
                employee.payroll_emp_id
            )
            .first()
        )

        if payee:

            employee_ids = [payee.id]

            for dep in payee.payroll_dependents:
                employee_ids.append(dep.id)

            employee = payee

    entries = (
        db.query(TimeEntry)
        .filter(
            TimeEntry.employee_id.in_(employee_ids),
            TimeEntry.clock_in_at >= pp.start_at,
            TimeEntry.clock_in_at <= pp.end_at,
        )
        .order_by(TimeEntry.clock_in_at)
        .all()
    )

    rows = []

    # ==================================
    # Daily calculation
    # ==================================

    for te in entries:

        break_hours = 0
        start_break = None
        stop_break = None

        break_id = None

        if te.breaks:

            br = te.breaks[0]

            break_id = br.id

            start_break = br.start_at
            stop_break = br.end_at

            if br.start_at and br.end_at:

                break_hours = (
                    br.end_at - br.start_at
                ).total_seconds() / 3600

        total_hours = 0

        if te.clock_in_at and te.clock_out_at:

            total_hours = (
                te.clock_out_at -
                te.clock_in_at
            ).total_seconds() / 3600

            total_hours -= break_hours

        reg_hours = min(8, total_hours)

        ot_hours = max(
            0,
            total_hours - 8
        )

        rows.append({
            "id": te.id,

            "break_id": break_id,

            "date":
                te.clock_in_at.date(),

            "clock_in":
                te.clock_in_at,

            "clock_out":
                te.clock_out_at,

            "start_break":
                start_break,

            "stop_break":
                stop_break,

            "break_hours":
                round(break_hours, 2),

            "total_hours":
                round(total_hours, 2),

            "reg_hours":
                round(reg_hours, 2),

            "ot_hours":
                round(ot_hours, 2),

            "note":
                te.notes or "",

            "six_day_ot":
                False,
            "employee_name": te.employee.name if te.employee else "",
        })

    # ==================================
    # Six day rule
    # ==================================

    weeks = defaultdict(list)

    for row in rows:

        wk = week_key(
            datetime.combine(
                row["date"],
                datetime.min.time()
            )
        )

        weeks[wk].append(row)

    for week_rows in weeks.values():

        worked = [
            r
            for r in week_rows
            if r["total_hours"] > 0
        ]

        if len(worked) >= 6:

            lowest = min(
                worked,
                key=lambda x: x["total_hours"]
            )

            lowest["six_day_ot"] = True

            lowest["ot_hours"] += lowest["reg_hours"]

            lowest["reg_hours"] = 0

    # ==================================
    # Pay calculation
    # ==================================

    sum_break = 0
    sum_reg = 0
    sum_ot = 0

    sum_pay_reg = 0
    sum_pay_ot = 0

   

    for row in rows:

        pay_reg = (
            row["reg_hours"] * rate
        )

        pay_ot = (
            row["ot_hours"] * ot_rate
        )

        pay_total = (
            pay_reg + pay_ot
        )

        row["rate"] = rate
        row["ot_rate"] = ot_rate

        row["pay_reg"] = round(
            pay_reg,
            2
        )

        row["pay_ot"] = round(
            pay_ot,
            2
        )

        row["pay_total"] = round(
            pay_total,
            2
        )

        sum_break += row["break_hours"]
        sum_reg += row["reg_hours"]
        sum_ot += row["ot_hours"]

        sum_pay_reg += pay_reg
        sum_pay_ot += pay_ot

    # Time leave 

    leave_rows = (
        db.query(TimeLeave)
        .filter(
            TimeLeave.employee_id.in_(employee_ids),
            # TimeLeave.status == "approved",
            # TimeLeave.end_at >= pp.start_at,
            # TimeLeave.start_at <= pp.end_at,
        )
        .order_by(TimeLeave.start_at)
        .all()
    )
    # print("leave_rows", leave_rows)
    leave_total = (
        db.query(
            func.coalesce(
                func.sum(TimeLeave.hours),
                0
            )
        )
        .filter(
            TimeLeave.employee_id.in_(employee_ids),
            # TimeLeave.status == "approved",
            # TimeLeave.end_at >= pp.start_at,
            # TimeLeave.start_at <= pp.end_at,
        )
        .scalar()
    )

    return {
        "employee": {
            "id": employee.id,
            "name": employee.name,
            "payroll_emp_id": employee.payroll_emp_id,
            "payroll_dependents": [
                {
                    "id": d.id,
                    "name": d.name,
                    "payroll_emp_id": d.payroll_emp_id,
                }
                for d in employee.payroll_dependents
            ]
        },

         "pay_period": {
            "start_at": pp.start_at,
            "end_at": pp.end_at,
        },


        "rows": rows,

        "total_break_hours": round(sum_break, 2),
        "total_reg_hours": round(sum_reg, 2),
        "total_ot_hours": round(sum_ot, 2),

        "total_pay_reg": round(sum_pay_reg, 2),
        "total_pay_ot": round(sum_pay_ot, 2),
        "total_pay": round(
            sum_pay_reg + sum_pay_ot,
            2
        ),

        "leave_rows": leave_rows,
        "total_leave_hours": float(leave_total),
    }

@router.get("/calculate")
def calculate_payroll(
    employee_id: int,
    pp_id: int,
    rate: float,
    ot_rate: float,
    show_payee: bool = False,
    db: Session = Depends(get_db),
):
    return calculate_payroll_data(
        employee_id,
        pp_id,
        rate,
        ot_rate,
        show_payee,
        db
    )
    
    
@router.get("/export_excel")
def export_excel(
    employee_id: int,
    pp_id: int,
    rate: float,
    ot_rate: float,
    show_payee: bool = False,
    db: Session = Depends(get_db),
):

    payroll = calculate_payroll(
        employee_id=employee_id,
        pp_id=pp_id,
        rate=rate,
        ot_rate=ot_rate,
        show_payee=show_payee,
        db=db
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"

    # ws["A1"] = "Employee"
    # ws["B1"] = payroll["employee"]["name"]

    # ws["A2"] = "Pay Period"
    # ws["B2"] = (
    #     f"{payroll['pay_period']['start_at']:%m/%d/%y}"
    #     f" - "
    #     f"{payroll['pay_period']['end_at']:%m/%d/%y}"
    # )

    # ==========================
    # Header
    # ==========================


    ws["A1"] = "No"
    ws["B1"] = "Date"
    ws["C1"] = "Clock In"
    ws["D1"] = "Start Break"
    ws["E1"] = "Stop Break"
    ws["F1"] = "Clock Out"
    ws["G1"] = "Break Hours"
    ws["H1"] = "Reg Hours"
    ws["I1"] = "OT Hours"
    ws["J1"] = "Rate"
    ws["K1"] = "OT Rate"
    ws["L1"] = "Pay Reg"
    ws["M1"] = "Pay OT"
    ws["N1"] = "Total Pay"
    ws["O1"] = "Note"

    for cell in ws[1]:
        cell.font = Font(
            bold=True
        )
    # ==========================
    # Detail
    # ==========================

    row_no = 2

    for idx, r in enumerate(payroll["rows"], start=1):

      
        ws.cell(row_no, 1, idx)

        ws.cell(
            row_no,
            2,
            r["date"].strftime("%m/%d/%y")
            if r["date"] else ""
        )

        ws.cell(
            row_no,
            3,
            r["clock_in"].strftime("%H:%M")
            if r["clock_in"] else ""
        )

        ws.cell(
            row_no,
            4,
            r["start_break"].strftime("%H:%M")
            if r["start_break"] else ""
        )

        ws.cell(
            row_no,
            5,
            r["stop_break"].strftime("%H:%M")
            if r["stop_break"] else ""
        )

        ws.cell(
            row_no,
            6,
            r["clock_out"].strftime("%H:%M")
            if r["clock_out"] else ""
        )

        ws.cell(row_no, 7, r["break_hours"])
        ws.cell(row_no, 8, r["reg_hours"])
        ws.cell(row_no, 9, r["ot_hours"])

        for col in [7, 8, 9]:
            ws.cell(
                row_no,
                col
            ).number_format = '0.00'

        ws.cell(row_no, 10, r["rate"])
        ws.cell(row_no, 11, r["ot_rate"])

        ws.cell(row_no, 12, r["pay_reg"])
        ws.cell(row_no, 13, r["pay_ot"])
        ws.cell(row_no, 14, r["pay_total"])
        

        for col in [10, 11, 12, 13, 14]:
            ws.cell(
                row_no,
                col
            ).number_format = '$#,##0.00'

        ws.cell(row_no, 15, r["note"])

        row_no += 1
        


    # ==========================
    # Total
    # ==========================

    ws.cell(row_no, 7, "Total")

    ws.cell(
        row_no,
        8,
        payroll["total_reg_hours"]
    )

    ws.cell(
        row_no,
        9,
        payroll["total_ot_hours"]
    )

    for col in [8, 9]:
        ws.cell(
            row_no,
            col
        ).number_format = '0.00'

    ws.cell(
        row_no,
        12,
        payroll["total_pay_reg"]
    )

    ws.cell(
        row_no,
        13,
        payroll["total_pay_ot"]
    )

    ws.cell(
        row_no,
        14,
        payroll["total_pay"]
    )

    for col in [12, 13, 14]:
        ws.cell(
            row_no,
            col
        ).number_format = '$#,##0.00'


    ### Time leave 
    row_no += 3
    
    ws.cell(row_no, 1, "No")
    ws.cell(row_no, 2, "Start Date")
    ws.cell(row_no, 3, "End Date")
    ws.cell(row_no, 4, "Type")
    ws.cell(row_no, 5, "Hours")

    row_no += 1

    for idx, leave in enumerate(
        payroll["leave_rows"],
        start=1
    ):

        ws.cell(
            row_no,
            1,
            idx
        )

        ws.cell(
            row_no,
            2,
            leave.start_at.strftime("%m/%d/%y")
        )

        ws.cell(
            row_no,
            3,
            leave.end_at.strftime("%m/%d/%y")
        )

        ws.cell(
            row_no,
            4,
            leave.leave_type
        )

        ws.cell(
            row_no,
            5,
            float(leave.hours or 0)
        )

        row_no += 1

    ws.cell(row_no, 4, "Total Leave")
    ws.cell(
        row_no,
        5,
        payroll["total_leave_hours"]
    )
    # ==========================
    # Save Memory Stream
    # ==========================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    start_date = payroll["pay_period"]["start_at"].strftime("%Y-%m-%d")
    end_date = payroll["pay_period"]["end_at"].strftime("%Y-%m-%d")
    filename = (
        f"{payroll['employee']['name']}_{start_date}_{end_date}.xlsx"
    )

    # filename = (
    #     f"Payroll_"
    #     f"{payroll['employee']['name']}.xlsx"
    # )

    return StreamingResponse(
        output,
        media_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f'attachment; filename="{filename}"'
        }
    )

@router.get("/export_payee")
def export_payee(
    employee_id: int,
    pp_id: int,
    rate: float,
    ot_rate: float,
    db: Session = Depends(get_db),
):

    payroll = calculate_payroll_data(
        employee_id=employee_id,
        pp_id=pp_id,
        rate=rate,
        ot_rate=ot_rate,
        show_payee=True,
        db=db
    )

    wb = Workbook()
    ws = wb.active
    ws.title = ( f"{payroll['employee']['name']}"   
)


    # ==========================
    # Detail Header
    # ==========================

    ws["A1"] = "No"
    ws["B1"] = "Date"
    ws["C1"] = "Clock In"
    ws["D1"] = "Start Break"
    ws["E1"] = "Stop Break"
    ws["F1"] = "Clock Out"
    ws["G1"] = "Break Hours"
    ws["H1"] = "Reg Hours"
    ws["I1"] = "OT Hours"
    ws["J1"] = "Rate"
    ws["K1"] = "OT Rate"
    ws["L1"] = "Pay Reg"
    ws["M1"] = "Pay OT"
    ws["N1"] = "Total Pay"
    ws["O1"] = "Employee"

    row_no = 2

    for idx, r in enumerate(payroll["rows"], start=1):

        ws.cell(row_no, 1, idx)
        ws.cell(row_no, 2, r["date"].strftime("%m/%d/%y"))
        ws.cell(row_no, 3, r["clock_in"].strftime("%H:%M") if r["clock_in"] else "")
        ws.cell(row_no, 4, r["start_break"].strftime("%H:%M") if r["start_break"] else "")
        ws.cell(row_no, 5, r["stop_break"].strftime("%H:%M") if r["stop_break"] else "")
        ws.cell(row_no, 6, r["clock_out"].strftime("%H:%M") if r["clock_out"] else "")

        ws.cell(row_no, 7, r["break_hours"])
        ws.cell(row_no, 8, r["reg_hours"])
        ws.cell(row_no, 9, r["ot_hours"])

        for col in [7, 8, 9]:
            ws.cell(
                row_no,
                col
            ).number_format = '0.00'

        ws.cell(row_no, 10, r["rate"])
        ws.cell(row_no, 11, r["ot_rate"])

        ws.cell(row_no, 12, r["pay_reg"])
        ws.cell(row_no, 13, r["pay_ot"])
        ws.cell(row_no, 14, r["pay_total"])
        ws.cell( row_no, 15, r.get("employee_name", "")        )
        row_no += 1

    row_no += 1
    
    # ==========================
    # Total
    # ==========================

    ws.cell(row_no, 7, "Total")

    ws.cell(
        row_no,
        8,
        payroll["total_reg_hours"]
    )

    ws.cell(
        row_no,
        9,
        payroll["total_ot_hours"]
    )

    ws.cell(
        row_no,
        12,
        payroll["total_pay_reg"]
    )

    for col in [8, 9]:
        ws.cell(
            row_no,
            col
        ).number_format = '0.00'

    ws.cell(
        row_no,
        13,
        payroll["total_pay_ot"]
    )

    ws.cell(
        row_no,
        14,
        payroll["total_pay"]
    )

    # Format $
    for col in [12, 13, 14]:
        ws.cell(
            row_no,
            col
        ).number_format = '$#,##0.00'

    # ขีดหนา Total
    for col in range(7, 15):
        ws.cell(
            row_no,
            col
        ).font = Font(bold=True)

    row_no += 2

    # # ==========================
    # # Dependents
    # # ==========================


    # for dep in payroll["employee"]["payroll_dependents"]:
    #     if dep["name"] == payroll["employee"]["name"]:
    #         continue

    #     dep_payroll = calculate_payroll_data(
    #         employee_id=dep["id"],
    #         pp_id=pp_id,
    #         rate=rate,
    #         ot_rate=ot_rate,
    #         show_payee=False,
    #         db=db
    #     )

    #     ws.cell(row_no, 12, dep["name"])

    #     ws.cell(
    #         row_no,
    #         14,
    #         dep_payroll["total_pay"]
    #     )

    #     ws.cell(
    #         row_no,
    #         14
    #     ).number_format = '$#,##0.00'
    #     row_no += 1
    # row_no += 2

    # ==========================
    # Payee Pay
    # ==========================

    payee_payroll = calculate_payroll_data(
        employee_id=payroll["employee"]["id"],
        pp_id=pp_id,
        rate=rate,
        ot_rate=ot_rate,
        show_payee=False,
        db=db
    )

    payee_total = payroll["total_pay"]

    # ==========================
    # Dependent Pay
    # ==========================

    employee_totals = []
    employee_sum = 0
    payee_name = payroll["employee"]["name"]
    for dep in payroll["employee"]["payroll_dependents"]:
        if dep["name"] == payee_name:
         continue

        dep_payroll = calculate_payroll_data(
            employee_id=dep["id"],
            pp_id=pp_id,
            rate=rate,
            ot_rate=ot_rate,
            show_payee=False,
            db=db
        )

        employee_totals.append({
            "id": dep["id"],
            "name": dep["name"],
            "total_pay": dep_payroll["total_pay"]
        })
        # print("employee_sum + dep_payroll[total_pay]: " , employee_sum, dep_payroll["total_p/ay"])
        employee_sum += dep_payroll["total_pay"]

    ws.cell(row_no, 12, "Dependent Pay")
    row_no += 1

    for emp in employee_totals:

        ws.cell(row_no, 12, emp["name"])

        ws.cell(
            row_no,
            14,
            emp["total_pay"]
        )

        ws.cell(row_no, 14).number_format = '$#,##0.00'

        row_no += 1

    ws.cell(row_no, 12,"Total")
    ws.cell(row_no, 14,employee_sum )
    ws.cell(row_no, 14).number_format = '$#,##0.00'
    row_no += 1

    # ==========================
    # Extra
    # ==========================
    # print("payee_total - employee_sum:",payee_total,employee_sum)
    extra_amount = payee_total - employee_sum

    row_no += 1

    ws.cell(row_no, 12, "Extra")
    ws.cell(row_no, 14, extra_amount)
    ws.cell(row_no, 14).number_format = '$#,##0.00'

    # ==========================
    # As OT Hours
    # ==========================

    row_no += 1

    div_ot = (
        extra_amount / ot_rate
        if ot_rate else 0
    )

    ws.cell(row_no, 12, "As OT Hours")
    ws.cell(row_no, 14, round(div_ot, 2))

    # ==========================
    # OT Allocation
    # ==========================

    dep_count = len(employee_totals)

    split_team = (
        div_ot / dep_count
        if dep_count else 0
    )

    ot_paid = split_team * ot_rate

    row_no += 2


    ot_paid_all = 0

    for emp in employee_totals:

        ot_paid_all += ot_paid

        ws.cell(row_no, 12, emp["name"])

        ws.cell(
            row_no,
            13,
            round(split_team, 2)
        )

        ws.cell(
            row_no,
            14,
            round(ot_paid, 2)
        )

        ws.cell(
            row_no,
            14
        ).number_format = '$#,##0.00'

        row_no += 1

    ws.cell(row_no, 12, "Extra")
    ws.cell(row_no, 14, round(ot_paid_all, 2))
    ws.cell(row_no, 14).number_format = '$#,##0.00'

    # ==========================
    # Final Pay
    # ==========================

    row_no += 2

    ws.cell(row_no, 12, "Final Pay")
    row_no += 1

    grand_total = 0

    for emp in employee_totals:

        final_pay = emp["total_pay"] + ot_paid

        grand_total += final_pay

        ws.cell(row_no, 12, emp["name"])

        ws.cell(
            row_no,
            14,
            round(final_pay, 2)
        )

        ws.cell(
            row_no,
            14
        ).number_format = '$#,##0.00'

        row_no += 1

    ws.cell(row_no, 12, "Total")
    ws.cell(row_no, 14, round(grand_total, 2))
    ws.cell(row_no, 14).number_format = '$#,##0.00'

    # ==========================
    # Time Leave
    # ==========================

    row_no += 3

    ws.cell(row_no, 1, "No")
    ws.cell(row_no, 2, "Start Date")
    ws.cell(row_no, 3, "End Date")
    ws.cell(row_no, 4, "Type")
    ws.cell(row_no, 5, "Hours")

    for col in range(1, 6):
        ws.cell(row_no, col).font = Font(bold=True)

    row_no += 1

    for idx, leave in enumerate(
        payroll["leave_rows"],
        start=1
    ):

        ws.cell(row_no, 1, idx)

        ws.cell(
            row_no,
            2,
            leave.start_at.strftime("%m/%d/%y")
        )

        ws.cell(
            row_no,
            3,
            leave.end_at.strftime("%m/%d/%y")
        )

        ws.cell(
            row_no,
            4,
            leave.leave_type
        )

        ws.cell(
            row_no,
            5,
            float(leave.hours or 0)
        )

        row_no += 1

    ws.cell(row_no, 4, "Total Leave")

    ws.cell(
        row_no,
        5,
        payroll["total_leave_hours"]
    )

    ws.cell(row_no, 4).font = Font(bold=True)
    ws.cell(row_no, 5).font = Font(bold=True)

    # ==========================
    # Export
    # ==========================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    start_date = payroll["pay_period"]["start_at"].strftime("%Y-%m-%d")
    end_date = payroll["pay_period"]["end_at"].strftime("%Y-%m-%d")

    filename = (
        f"{payroll['employee']['name']}_{start_date}_{end_date}.xlsx"
    )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f'attachment; filename="{filename}"'
        }
    )