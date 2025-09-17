# show_excel_data.py
# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
import re

def _to_str(x):
    return "" if x is None else str(x).strip()

def find_label_value(ws, labels, max_rows=30, max_cols=50):
    labset = {l.strip().lower() for l in labels}
    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if str(v).strip().lower() in labset:
                for cc in range(c + 1, min(c + 12, max_cols + 1)):
                    nv = ws.cell(r, cc).value
                    s = _to_str(nv)
                    if s:
                        return s
                return ""
    return ""

def read_metadata(path, sheet):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]

    part_no   = find_label_value(ws, {"part no.", "part no"})
    part_name = find_label_value(ws, {"part name.", "part name"})
    customer  = find_label_value(ws, {"customer.", "customer"})
    rev       = find_label_value(ws, {"rev.", "rev"})

    if not customer:
        # fallback หา pattern เช่น AF6182
        for r in range(1, 15):
            for c in range(1, 40):
                s = _to_str(ws.cell(r, c).value)
                if re.fullmatch(r"[A-Za-z]{1,3}\d{3,6}", s):
                    customer = s
                    break
            if customer:
                break

    return {
        "part_no": part_no or None,
        "part_name": part_name or None,
        "customer": customer or None,
        "rev": rev or None,
    }

# หาหัวตาราง
KEYS = {"Lot Number", "PO Number", "PO Date", "Qty PO", "Shipped / Date", "Qty Shipped"}
def find_header_row(path, sheet):
    raw = pd.read_excel(path, sheet_name=sheet, header=None, engine="openpyxl")
    for i in range(0, min(40, len(raw))):
        vals = set(_to_str(v) for v in raw.iloc[i].tolist())
        if len(KEYS & vals) >= 3:
            return i
    return 0

def read_table(path, sheet):
    hdr = find_header_row(path, sheet)
    df  = pd.read_excel(path, sheet_name=sheet, header=hdr, engine="openpyxl")
    df.columns = [_to_str(c) for c in df.columns]
    return df

def clean_table(df):
    want = [c for c in ["Lot Number","PO Number","PO Date","Qty PO","Shipped / Date","Qty Shipped"] if c in df.columns]
    df = df[want].copy()
    for d in ["PO Date", "Shipped / Date"]:
        if d in df.columns:
            df[d] = pd.to_datetime(df[d], errors="coerce")
    for n in ["Qty PO", "Qty Shipped"]:
        if n in df.columns:
            df[n] = pd.to_numeric(df[n], errors="coerce")
    df = df.dropna(how="all", subset=want).reset_index(drop=True)
    return df

if __name__ == "__main__":
    FILES = [
        (r"C:/Users/Tanapon/Downloads/2040364-1.xlsm", "2040364-1"),
        (r"C:/Users/Tanapon/Downloads/5673-22-1.xlsm", "5673-22-1"),
    ]

    for path, sheet in FILES:
        print("\n==============================")
        print(f"File: {path} [{sheet}]")
        meta = read_metadata(path, sheet)
        print("Metadata:", meta)

        df = clean_table(read_table(path, sheet))
        print("\n=== Columns ===")
        print(df.columns.tolist())
        print("\n=== Preview Data (10 rows) ===")
        print(df.head(10).to_string(index=False))
