
import sys, os

# เพิ่ม path ของ project ให้ Python มองเห็น
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from sqlalchemy import create_engine, text
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

DATABASE_URL = "postgresql+psycopg2://postgres:1234@100.88.56.126:5432/mydb"

engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)
db = SessionLocal()

from sqlalchemy.orm import Session
from models import Part, PartRevision, ProductionLot, PO, Customer
from sqlalchemy import desc

def get_part_info(db: Session, part_no: str):
    """
    Return (part_name, rev, customer_code)
    """
    # 1) หา part
    part = db.query(Part).filter(Part.part_no == part_no).first()
    if not part:
        return None, None, None
    
    # 2) หา rev ล่าสุด (is_current=True) ถ้าไม่มีก็เอาตัวล่าสุด
    rev_obj = (
        db.query(PartRevision)
        .filter(PartRevision.part_id == part.id, PartRevision.is_current == True)
        .order_by(desc(PartRevision.id))
        .first()
    )
    if rev_obj:
        rev = rev_obj.rev
    else:
        rev = None

    # 3) หาลูกค้าที่เคยมี LOT ล่าสุดของ part นี้
    latest_lot = (
        db.query(ProductionLot)
        .filter(ProductionLot.part_id == part.id)
        .order_by(desc(ProductionLot.id))
        .first()
    )

    customer_code = None
    if latest_lot and latest_lot.po_id:
        po = db.query(PO).filter(PO.id == latest_lot.po_id).first()
        if po:
            customer = db.query(Customer).filter(Customer.id == po.customer_id).first()
            if customer:
                customer_code = customer.code

    return part.name, rev, customer_code

from openpyxl import load_workbook
import os
import copy
import shutil


# ===============================
# CONFIG
# ===============================
SOURCE_FILE = r"C:\Data Base & Inventory Stock\source.xlsx"
DEST_FOLDER = r"C:\Data Base & Inventory Stock\data"
TEMPLATE_FILE = r"C:\Data Base & Inventory Stock\Template form.xlsm"

START_ROW = 7
COLUMN_LOT = "B"
COLUMN_PO = "C"


# ===============================
# FUNCTION: Clone Row (copy style + value)
# ===============================
def clone_row(ws, source_row, target_row):
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        sc = ws.cell(row=source_row, column=col)
        tc = ws.cell(row=target_row, column=col)

        tc.value = sc.value

        if sc.has_style:
            tc._style = copy.copy(sc._style)

        tc.number_format = sc.number_format

        if sc.hyperlink:
            tc._hyperlink = copy.copy(sc.hyperlink)

        if sc.comment:
            tc.comment = copy.copy(sc.comment)



# ===============================
# MAIN PROCESS
# ===============================
src_wb = load_workbook(SOURCE_FILE)
src = src_wb.active

for row in src.iter_rows(min_row=2, values_only=True):
    lot_number, po_number, part_no = row

    if not part_no:
        continue

    dest_file = os.path.join(DEST_FOLDER, f"{part_no}.xlsm")

    # If file missing → create from template
    if not os.path.exists(dest_file):
        print(f"📄 ไม่พบไฟล์ {dest_file} → สร้างใหม่จาก Template")
        shutil.copy(TEMPLATE_FILE, dest_file)

    print(f"\n📂 เปิดไฟล์: {dest_file}")

    dst_wb = load_workbook(dest_file, keep_vba=True)
    ws = dst_wb.active


    # =============================
    # STEP 1: Set HEADER (Part No, Name, Customer, Rev)
    # =============================
    part_name, rev, customer = get_part_info(db, part_no)
    ws["C2"] = part_no      # Part No
    ws["F2"] = part_name           # Part Name (Tony มีข้อมูลไหม?)
    ws["J2"] = customer           # Customer
    ws["L2"] = rev          # Rev

    print(f"   ✔ ตั้งค่า C2 PartNo = {part_no}")
    print(f"   ✔ Part Header: {part_no}  Name={part_name}  Rev={rev}  Customer={customer}")

    # =============================
    # STEP 2: Check duplicate LOT
    # =============================
    duplicate = False
    for r in range(START_ROW, ws.max_row + 1):
        if ws[f"{COLUMN_LOT}{r}"].value == lot_number:
            duplicate = True
            print(f"⛔ LOT ซ้ำ: {lot_number} → ข้าม")
            break

    if duplicate:
        dst_wb.save(dest_file)
        continue


    # =============================
    # STEP 3: หาแถวว่าง
    # =============================
    last_row = ws.max_row
    insert_row = None

    for r in range(START_ROW, last_row + 1):
        if ws[f"{COLUMN_LOT}{r}"].value in (None, ""):
            insert_row = r
            break

    if not insert_row:
        insert_row = last_row + 1


    # =============================
    # STEP 4: ถ้าไม่ว่าง → clone row ลงล่าง
    # =============================
    # row_empty = all(ws.cell(row=insert_row, column=c).value in (None, "")
    #                 for c in range(1, ws.max_column + 1))

    # if not row_empty:
    #     print(f"⚠ Row {insert_row} มีข้อมูล → clone ลงล่างแทน insert !!!!!!!!!!")
    #     clone_row(ws, insert_row, insert_row + 1)
    # else:
    #     print(f"   ✔ Row {insert_row} ว่าง พร้อม insert >>>>>>>")

    # =============================
# STEP 4: ถ้าไม่ว่าง → clone row ลงล่างจนกว่าจะว่าง
# =============================
    # 1) หาแถวที่จะ insert = 1 แถวล่างสุดของ sheet
    insert_row = ws.max_row + 1

    # 2) insert แถวว่างจริงใน excel
    ws.insert_rows(insert_row)

    # 3) clone template row ลงไป (เช่น row 8 เป็น template)
    TEMPLATE_ROW = 8    # แถวแม่แบบที่จัดสี/format ไว้
    clone_row(ws, TEMPLATE_ROW, insert_row)

    # 4) เคลียร์ค่าที่ clone มาจาก template แต่เก็บ format
    for c in range(1, ws.max_column + 1):
        ws.cell(row=insert_row, column=c).value = None


    # =============================
    # STEP 5: Insert LOT + PO
    # =============================
    ws[f"{COLUMN_LOT}{insert_row}"] = lot_number
    ws[f"{COLUMN_PO}{insert_row}"] = po_number

    print(f"   ✔ Inserted LOT={lot_number}, PO={po_number} → row {insert_row}")


    dst_wb.save(dest_file)

print("\n🎉 DONE — Header + Insert LOT + Auto Template + No Duplicates")
