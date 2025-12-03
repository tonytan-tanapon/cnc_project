
import sys, os

# เพิ่ม path ของ project ให้ Python มองเห็น
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from sqlalchemy import create_engine, text
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

DATABASE_URL = "postgresql+psycopg2://postgres:1234@100.88.56.126:5432/mydb"

engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)
db = SessionLocal()

from sqlalchemy.orm import Session
from models import Part, PartRevision, ProductionLot, PO, Customer
from sqlalchemy import desc

def get_part_info(db: Session, part_no: str):
    """
    Return (part_name, rev, customer_code)
    """
    # 1) หา part
    part = db.query(Part).filter(Part.part_no == part_no).first()
    if not part:
        return None, None, None
    
    # 2) หา rev ล่าสุด (is_current=True) ถ้าไม่มีก็เอาตัวล่าสุด
    rev_obj = (
        db.query(PartRevision)
        .filter(PartRevision.part_id == part.id, PartRevision.is_current == True)
        .order_by(desc(PartRevision.id))
        .first()
    )
    if rev_obj:
        rev = rev_obj.rev
    else:
        rev = None

    # 3) หาลูกค้าที่เคยมี LOT ล่าสุดของ part นี้
    latest_lot = (
        db.query(ProductionLot)
        .filter(ProductionLot.part_id == part.id)
        .order_by(desc(ProductionLot.id))
        .first()
    )

    customer_code = None
    if latest_lot and latest_lot.po_id:
        po = db.query(PO).filter(PO.id == latest_lot.po_id).first()
        if po:
            customer = db.query(Customer).filter(Customer.id == po.customer_id).first()
            if customer:
                customer_code = customer.code

    return part.name, rev, customer_code

from openpyxl import load_workbook
import os
import copy
import shutil


# ===============================
# CONFIG
# ===============================
# Folder of the current script (excel/)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# source.xlsx is inside the same folder
SOURCE_FILE = os.path.join(BASE_DIR, "source.xlsx")
DEST_FOLDER = r"Z:\Topnotch Group\Public\Data Base & Inventory Stock\data"
TEMPLATE_FILE = r"Z:\Topnotch Group\Public\Data Base & Inventory Stock\Template form.xlsm"



src_wb = load_workbook(SOURCE_FILE)
src = src_wb.active

for row in src.iter_rows(min_row=2, values_only=True):
    print(">> ", row)
    # lot_number, po_number, part_no = row
    date_po, name_cus,lot_number,po_number,part_no,part_desc, part_rev, duedate, qty_po,_ = row

# ##

# START_ROW = 7
# COLUMN_LOT = "B"
# COLUMN_PO = "C"
# COLUMN_PO_DATE = "E"
# COLUMN_QTY_PO = "F"
# COLUMN_DUEDATE = "G"



# # ===============================
# # FUNCTION: Clone Row (copy style + value)
# # ===============================
# def clone_row(ws, source_row, target_row):
#     max_col = ws.max_column

#     for col in range(1, max_col + 1):
#         sc = ws.cell(row=source_row, column=col)
#         tc = ws.cell(row=target_row, column=col)

#         tc.value = sc.value

#         if sc.has_style:
#             tc._style = copy.copy(sc._style)

#         tc.number_format = sc.number_format

#         if sc.hyperlink:
#             tc._hyperlink = copy.copy(sc.hyperlink)

#         if sc.comment:
#             tc.comment = copy.copy(sc.comment)



# # ===============================
# # MAIN PROCESS
# # ===============================
# src_wb = load_workbook(SOURCE_FILE)
# src = src_wb.active

# for row in src.iter_rows(min_row=2, values_only=True):
#     print(">> ", row)
#     # lot_number, po_number, part_no = row
#     date_po, name_cus,lot_number,po_number,part_no,part_desc, part_rev, duedate, qty_po,_ = row


#     if not part_no:
#         continue

#     dest_file = os.path.join(DEST_FOLDER, f"{part_no}.xlsm")

#     # If file missing → create from template
#     if not os.path.exists(dest_file):
#         print(f"📄 ไม่พบไฟล์ {dest_file} → สร้างใหม่จาก Template")
#         shutil.copy(TEMPLATE_FILE, dest_file)

#     print(f"\n📂 เปิดไฟล์: {dest_file}")

#     dst_wb = load_workbook(dest_file, keep_vba=True)
#     ws = dst_wb.active


#     # =============================
#     # STEP 1: Set HEADER (Part No, Name, Customer, Rev)
#     # =============================
#     part_name, rev, customer = get_part_info(db, part_no)
#     ws["C2"] = part_no      # Part No
#     ws["F2"] = part_name           # Part Name (Tony มีข้อมูลไหม?)
#     ws["J2"] = customer           # Customer
#     ws["L2"] = rev          # Rev

#     print(f"   ✔ ตั้งค่า C2 PartNo = {part_no}")
#     print(f"   ✔ Part Header: {part_no}  Name={part_name}  Rev={rev}  Customer={customer}")

#     # =============================
#     # STEP 2: Check duplicate LOT
#     # =============================
#     duplicate = False
#     for r in range(START_ROW, ws.max_row + 1):
#         if ws[f"{COLUMN_LOT}{r}"].value == lot_number:
#             duplicate = True
#             print(f"⛔ LOT ซ้ำ: {lot_number} → ข้าม")
#             break

#     if duplicate:
#         dst_wb.save(dest_file)
#         continue


#     # =============================
#     # STEP 3: หาแถว LOT สุดท้าย
#     # =============================
#     last_lot_row = None
#     for r in range(START_ROW, ws.max_row + 1):
#         val = ws[f"{COLUMN_LOT}{r}"].value
#         if val not in (None, ""):
#             last_lot_row = r

#     if last_lot_row is None:
#         # ยังไม่มี LOT เลย → แถวแรกที่จะใช้ = START_ROW
#         insert_row = START_ROW
#     else:
#         # มี LOT แล้ว → แทรกแถวใหม่ใต้ LOT สุดท้าย
#         insert_row = last_lot_row + 1

#     # =============================
#     # STEP 4: Insert LOT + PO
#     # =============================
#     ws[f"{COLUMN_LOT}{insert_row}"] = lot_number
#     ws[f"{COLUMN_PO}{insert_row}"] = po_number
#     ws[f"{COLUMN_PO_DATE}{insert_row}"] = date_po
#     ws[f"{COLUMN_QTY_PO}{insert_row}"] = qty_po

#     from dateutil.relativedelta import relativedelta
#     import datetime

#     if duedate is not None:
#         one_month_ago = duedate - relativedelta(months=1)

#         # weekday(): Monday=0 ... Sunday=6
#         weekday = one_month_ago.weekday()

#         if weekday == 5:          # Saturday
#             # Move forward to next Friday (6 days later)
#             one_month_ago += datetime.timedelta(days=6)
#         elif weekday == 6:        # Sunday
#             # Move forward to next Friday (5 days later)
#             one_month_ago += datetime.timedelta(days=5)

#         ws[f"{COLUMN_DUEDATE}{insert_row}"] = one_month_ago

   
#     # =============================
#     # STEP 5 : Cascade shift ลงล่างสำหรับ column อื่น ๆ
#     # =============================

#     # =============================
#     # STEP 5 : Column-wise shift แบบ temp buffer
#     # =============================

#     COLUMNS_TO_MOVE = list(range(9, 11))  # D–K = 4..11
#     row = insert_row
#     next_row = insert_row + 1

#     print(f"   🔄 Column Shift แบบ temp buffer row={row} → row={next_row}")

#     for col in COLUMNS_TO_MOVE:

#         # 1) ค่าเริ่มต้นที่ row,col
#         src_val = ws.cell(row=row, column=col).value

#         # ถ้า row ว่าง → ไม่มีอะไรจะเลื่อน ไป col ถัดไป
#         if src_val in (None, ""):
#             continue

#         temp = src_val   # ค่าแรกที่จะเลื่อนลงไปแถวถัดไป
#         r = next_row

#         while True:
#             cell = ws.cell(r, col)

#             # 2) ถ้าแถวนี้ว่าง → วางแล้วจบ
#             if cell.value in (None, ""):
#                 cell.value = temp
#                 break

#             # 3) ถ้าแถวนี้มีข้อมูล → เลื่อนข้อมูลลง
#             old = cell.value      # เก็บค่าที่จะโดนทับ
#             cell.value = temp     # วางค่าของบนลง
#             temp = old            # แล้ว temp = ค่าที่ขยับลง

#             # ไปแถวถัดไป
#             r += 1
    
#     # =============================
#     # CLEAR แถวแรกหลัง shift
#     # =============================
#     print(f"   🧹 Clear original row {row} (D–K)")

#     for col in COLUMNS_TO_MOVE:
#         cell = ws.cell(row=row, column=col)
#         if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
#             continue
#         cell.value = None

#     dst_wb.save(dest_file)

# print("\n🎉 DONE — Header + Insert LOT + Auto Template + No Duplicates")
