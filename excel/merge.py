import os
import csv
from openpyxl import load_workbook
from concurrent.futures import ProcessPoolExecutor, as_completed

SOURCE_FOLDER = r"C:\Data Base & Inventory Stock\data"
SOURCE_FOLDER_EXCEL = r"C:\Data Base & Inventory Stock\data_excel"
DEST_FILE = r"C:\Data Base & Inventory Stock\lot_export.csv"


def extract_lot_rows(file_path):
    try:
        wb = load_workbook(file_path, data_only=True, keep_vba=True)
        ws = wb.active

        rows = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            lot_no = row[1]  # Column B
            if lot_no and isinstance(lot_no, str) and lot_no.startswith("L"):
                rows.append(row)

        return rows

    except Exception as e:
        return f"ERROR::{file_path}::{e}"


def process_all_files_parallel(folder, output_file):

    header = [
        "No", "Lot Number", "PO Number", "Prod Qty", "PO Date",
        "Qty PO", "Due Date", "Qty Shipped", "First Article",
        "Remark", "Tracking No", "Real Shipped Date", "Incoming Stock",
        "Received QTY", "Name Inspection", "Remark (QA Inspection)",
        "Rework/Repair", "*Remark (Rework)", "Qty Reject", "*Remark (Reject)",
        "Incoming Rework", "Finish goods in stock", "Lot Number", "PO Number",
        "Qty Take Out", "Date Take Out Stock",
        "empty",
        "WIP", "WIP Cont w/Lot", "QTY Rework", "Green Tag N.",
        "Rework w/Lot", "QTY Prod", "QTY Shipped", "Residual",
        "QTY Use", "Balance", "Scrap Later", "From Lot",
        "ST Status", "Note", "Date"
    ]

    # Filter files (skip temp files)
    files = [
        os.path.join(folder, f)
        for f in os.listdir(folder)
        if f.lower().endswith(".xlsm") and not f.startswith("~$")
    ]

    # Store all final rows
    all_rows = []

    # Parallel reading
    with ProcessPoolExecutor() as executor:
        futures = {executor.submit(extract_lot_rows, f): f for f in files}

        for future in as_completed(futures):
            result = future.result()

            if isinstance(result, str) and result.startswith("ERROR::"):
                print(result)
                continue

            for row in result:
                if row[0] != "No.":
                    all_rows.append(row)

    # Write CSV output
    # Write CSV output
    with open(output_file, "w", newline="", encoding="utf-8") as csvfile:

        # เลือกเฉพาะ Columns ที่ต้องการ
        selected_indexes = [1, 2, 3,4,5,6,7,8,9,10,11]  # Lot Number, PO Number, Prod Qty, Due Date
        selected_headers = [header[i] for i in selected_indexes]

        writer = csv.writer(csvfile)

        # เขียนหัวตารางแบบเลือกเฉพาะ col
        writer.writerow(selected_headers)

        for row in all_rows:
            # row เป็น tuple เช่น (No, Lot Number, PO Number, Prod Qty, ...)
            selected = [row[i] for i in selected_indexes]
            writer.writerow(selected)


    print("DONE! CSV exported to", output_file)


# ----------------------
# REQUIRED on Windows
# ----------------------
if __name__ == "__main__":
    process_all_files_parallel(SOURCE_FOLDER, DEST_FILE)
