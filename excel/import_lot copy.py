from openpyxl import load_workbook
import os
import copy

# ===============================
# CONFIG
# ===============================
SOURCE_FILE = r"C:\Data Base & Inventory Stock\source.xlsx"
DEST_FOLDER = r"C:\Data Base & Inventory Stock\data"

START_ROW = 7
COLUMN_LOT = "B"
COLUMN_PO = "C"


# ===============================
# FUNCTION: Clone Row (copy style + value)
# ===============================
def clone_row(ws, source_row, target_row):
    """Clone row ลง target_row แบบไม่เสีย format"""
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        sc = ws.cell(row=source_row, column=col)
        tc = ws.cell(row=target_row, column=col)

        # copy value
        tc.value = sc.value

        # copy style (important!)
        if sc.has_style:
            tc._style = copy.copy(sc._style)

        # number format
        tc.number_format = sc.number_format

        # hyperlink
        if sc.hyperlink:
            tc._hyperlink = copy.copy(sc.hyperlink)

        # comments
        if sc.comment:
            tc.comment = copy.copy(sc.comment)



# ===============================
# MAIN PROCESS
# ===============================
src_wb = load_workbook(SOURCE_FILE)
src = src_wb.active

for row in src.iter_rows(min_row=2, values_only=True):
    lot_number, po_number, part_no = row

    if not part_no:
        continue

    dest_file = os.path.join(DEST_FOLDER, f"{part_no}.xlsm")

    if not os.path.exists(dest_file):
        print(f"❌ ไม่พบไฟล์: {dest_file}")
        continue

    print(f"\n📂 เปิดไฟล์: {dest_file}")

    # load destination excel
    dst_wb = load_workbook(dest_file, keep_vba=True)
    ws = dst_wb.active

    # =============================
    # STEP 1: เช็ค LOT ซ้ำ
    # =============================
    duplicate = False
    for r in range(START_ROW, ws.max_row + 1):
        if ws[f"{COLUMN_LOT}{r}"].value == lot_number:
            duplicate = True
            print(f"⛔ LOT ซ้ำ: {lot_number} ในไฟล์ {part_no}.xlsm → ข้าม")
            break

    if duplicate:
        continue


    # =============================
    # STEP 2: หาแถวว่าง
    # =============================
    last_row = ws.max_row
    insert_row = None

    for r in range(START_ROW, last_row + 1):
        if ws[f"{COLUMN_LOT}{r}"].value in (None, ""):
            insert_row = r
            break

    if not insert_row:
        insert_row = last_row + 1


    # =============================
    # STEP 3: ถ้าแถวนั้นมีข้อมูล → clone row ลงล่าง
    # =============================
    row_empty = all(ws.cell(row=insert_row, column=c).value in (None, "")
                    for c in range(1, ws.max_column + 1))

    if not row_empty:
        print(f"⚠ Row {insert_row} มีข้อมูล → clone row ลงล่างแทน insert")
        clone_row(ws, insert_row, insert_row + 1)


    # =============================
    # STEP 4: Insert new data
    # =============================
    ws[f"{COLUMN_LOT}{insert_row}"] = lot_number
    ws[f"{COLUMN_PO}{insert_row}"] = po_number

    print(f"   ✔ Inserted LOT={lot_number}, PO={po_number} → row {insert_row}")


    # =============================
    # STEP 5: Save file
    # =============================
    dst_wb.save(dest_file)


print("\n🎉 DONE — ทุก part update แล้ว! Format ไม่พัง + กัน LOT ซ้ำ")
