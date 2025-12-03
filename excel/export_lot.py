import os
from openpyxl import load_workbook, Workbook

SOURCE_FOLDER = r"C:\Data Base & Inventory Stock\data"
DEST_FILE = r"C:\Data Base & Inventory Stock\lot_export.xlsx"

def extract_lot_rows(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active  # assume lot info is in the first sheet

    extracted_rows = []

    # Scan rows
    for row in ws.iter_rows(min_row=4, values_only=True):  # ignore header rows
       
        lot_no = row[1]  # Column B (index 1)

        if lot_no and isinstance(lot_no, str) and lot_no.startswith("L"):
            extracted_rows.append(row)

    return extracted_rows


def process_all_files(folder, output_file):
    master_wb = Workbook()
    master_ws = master_wb.active
    master_ws.title = "Lot Export"

    # Write header row
    header = [
        "No", "Lot Number", "PO Number", "Prod Qty", "PO Date",
        "Qty PO", "Due Date", "Qty Shipped", "First Article",
        "Remark", "Tracking No", "Real Shipped Date", "Incoming Stock",
        "Received QTY",	"Name Inspection",	"Remark (QA Inspection)",	"Rework/Repair",	"*Remark (Rework)",	"Qty Reject",	"*Remark (Reject)",
        "Incoming Rework",	"Finish goods in stock",	"Lot Number",	"PO Number",	"Qty Take Out",	"Date Take Out Stock",
        "empty",
        "WIP",	"WIP Cont w/Lot",	"QTY Rework",	"Green Tag N.",	"Rework w/Lot",	"QTY Prod",	"QTY Shipped",	"Residual", 	"QTY Use",	"Balance",	"Scrap Later",	"From Lot",	"ST Status",	"Note",	"Date",


					


						

    ]
    master_ws.append(header)

    for file in os.listdir(folder):
        if not file.lower().endswith(".xlsm"):
            continue

        file_path = os.path.join(folder, file)
        # print(f"Processing: {file_path}")

        lot_rows = extract_lot_rows(file_path)

        for row in lot_rows:
            # print("  Extracted:", row[0])
            if row[0] != "No.":
                master_ws.append(row)

    master_wb.save(output_file)
    print("DONE! Exported to", output_file)


process_all_files(SOURCE_FOLDER, DEST_FILE)
