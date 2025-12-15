## update lot from excel to database 
# get lot data from excel 
# update lot_qty, shipdate, tracking_no in database


import os
import csv
from openpyxl import load_workbook
from concurrent.futures import ProcessPoolExecutor, as_completed

SOURCE_FOLDER = r"Z:\Topnotch Group\Public\Data Base & Inventory Stock\Data"
# DEST_FILE = r"C:\Data Base & Inventory Stock\test\lot_export.csv"
DEST_FILE = ""
lot_data = {}

def extract_lot_rows(file_path):
    try:
        wb = load_workbook(file_path, data_only=True, keep_vba=True)
        ws = wb.active

        rows = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            lot_no = row[1]  # Column B
            if lot_no and isinstance(lot_no, str) and lot_no.startswith("L"):
                rows.append(row)

        return rows

    except Exception as e:
        return f"ERROR::{file_path}::{e}"

def update_to_database(part_name, row): 
    
    lot_no = row[1]
    prod_qty = row[3]
    due_date = row[6]
    qty_ship = row[7]
    tracking_no = row[10]
    ship_date = row[11] 
    take_out_qty = row[26]
    
    good_sotck = row[21]
    

    lot_data[lot_no]    = {
        "part_name": part_name,
        "prod_qty": prod_qty,
        "due_date": due_date,
        "qty_ship": qty_ship,
        "tracking_no": tracking_no,
        "ship_date": ship_date,
        
        "good_sotck": good_sotck,
       
    }
   
def process_all_files_parallel(folder, output_file):

    header = [
        "No", "Lot Number", "PO Number", "Prod Qty", "PO Date",
        "Qty PO", "Due Date", "Qty Shipped", "First Article",
        "Remark", "Tracking No", "Real Shipped Date", "Incoming Stock",
        "Received QTY", "Name Inspection", "Remark (QA Inspection)",
        "Rework/Repair", "*Remark (Rework)", "Qty Reject", "*Remark (Reject)",
        "Incoming Rework", "Finish goods in stock", "Lot Number", "PO Number",
        "Qty Take Out", "Date Take Out Stock",
        "empty",
        "WIP", "WIP Cont w/Lot", "QTY Rework", "Green Tag N.",
        "Rework w/Lot", "QTY Prod", "QTY Shipped", "Residual",
        "QTY Use", "Balance", "Scrap Later", "From Lot",
        "ST Status", "Note", "Date"
    ]

    # Filter files (skip temp files)
    files = [
        os.path.join(folder, f)
        for f in os.listdir(folder)
        if f.lower().endswith(".xlsm") and not f.startswith("~$")
    ]

    # Store all final rows
    all_rows = []
    
    # Parallel reading
    with ProcessPoolExecutor() as executor:
        futures = {executor.submit(extract_lot_rows, f): f for f in files}

        for future in as_completed(futures):
            filename = futures[future]  
            part_name = os.path.splitext(os.path.basename(filename))[0]
            result = future.result()

            if isinstance(result, str) and result.startswith("ERROR::"):
                print(result)
                continue

            for row in result:
                if row[0] != "No.":
                    # print(part_name,row[1:10])
                    update_to_database(part_name, row)
                    all_rows.append(row)
    
    

   
    return lot_data
    # Write CSV output
    # with open(output_file, "w", newline="", encoding="utf-8") as csvfile:
    #     writer = csv.writer(csvfile)
    #     writer.writerow(header)

    #     for row in all_rows:
            
    #         writer.writerow(row)

    # print("DONE! CSV exported to", output_file)


# ----------------------
# REQUIRED on Windows
# ----------------------
if __name__ == "__main__":
    process_all_files_parallel(SOURCE_FOLDER, DEST_FILE)
