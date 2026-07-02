#!/usr/bin/env python3
"""
Import material purchase data from CSV

CSV = RECEIVING (increase stock)
Allocate LotMaterialUse ONLY when stock exists
"""

from __future__ import annotations
from openpyxl import load_workbook
from decimal import Decimal
from datetime import datetime, date
from pathlib import Path
from zoneinfo import ZoneInfo
import sys, os

from sqlalchemy import create_engine, select, and_, func
from sqlalchemy.orm import Session, sessionmaker

# ---------- CONFIG ----------
LA_TZ = ZoneInfo("America/Los_Angeles")
DATABASE_URL = "postgresql+psycopg2://postgres:1234@localhost:5432/mydb"
# CSV_FILE = Path(r"C:\Users\TPSERVER\dev\cnc_project\database_import\import_material.csv")
# CSV_FILE = Path(r"C:\Users\TPSERVER\dev\cnc_project\database_import\import_material_mini.csv")
# CSV_FILE = Path(r"C:\Users\TPSERVER\dev\cnc_project\database_import\import_material_all.csv")
CSV_ENCODING = "utf-8-sig"



# ---------- IMPORT MODELS ----------
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from models import (
    Supplier,
    MaterialPO,
    MaterialPOLine,
    RawMaterial,
    RawBatch,
    Part,
    PartMaterial,
    ProductionLot,
    LotMaterialUse,
    LotMaterialUseHistory
)

# ---------- ENGINE ----------
engine = create_engine(DATABASE_URL, future=True)
SessionLocal = sessionmaker(bind=engine, class_=Session, autoflush=False, autocommit=False)

# =====================================================
# ================= HELPERS ===========================
# =====================================================

def normalize(s):
    return " ".join(str(s).strip().split()) if s else None


def parse_decimal(s):
    try:
        return Decimal(str(s).replace(",", "").replace("$", "").strip())
    except Exception:
        return None


def parse_date(s):
    if not s:
        return None
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except Exception:
            pass
    return None

# =====================================================
# ============= PO#, Qty PARSER =======================
# =====================================================

def parse_po_qty(raw_text):
    """
    Return list of {po, qty, note}
    """
    if not raw_text:
        return []

    records = []

    for line in raw_text.splitlines():
        line = line.strip().strip('"')
        if not line:
            continue

        # P123 & P456
        if "&" in line and "," not in line:
            for p in line.split("&"):
                records.append({"po": p.strip(), "qty": None, "note": "linked"})
            continue

        parts = [p.strip() for p in line.split(",")]
        po = parts[0]
        qty = 0
        note = None

        if len(parts) >= 2:
            if parts[1].lower() == "partial":
                note = "partial"
            else:
                try:
                    qty = Decimal(parts[1])
                except Exception:
                    note = parts[1]

        if len(parts) > 2:
            note = ",".join(parts[2:])

        records.append({"po": po, "qty": qty, "note": note})

    return records


# =====================================================
# ================= UPSERTS ===========================
# =====================================================

from datetime import date
from sqlalchemy import select

def upsert_supplier(db, name):

    code = str(name).strip().upper()

    s = db.scalar(
        select(Supplier).where(
            Supplier.code == code
        )
    )

    if s:
        s.name = name
        s.is_material_supplier = True
        return s

    s = Supplier(
        code=code,
        name=name,
        is_material_supplier=True,
    )

    db.add(s)
    db.flush()

    return s



def upsert_material_po(db, supplier, po_no, order_date):
    
    mpo = db.scalar(
        select(MaterialPO).where(MaterialPO.mat_po_no == po_no)
    )
    
    if mpo:
        
        # ✅ UPDATE existing PO
       
        mpo.supplier_id = supplier.id
        mpo.order_date = order_date or mpo.order_date or date.today()
        mpo.status = mpo.status or "open"
        return mpo

    # ✅ INSERT new PO
    mpo = MaterialPO(
       
        mat_po_no=po_no,
        supplier_id=supplier.id,
        order_date=order_date or date.today(),
        status="open",
    )
    db.add(mpo)
    db.flush()
    return mpo


from utils.code_generator import next_code

def upsert_raw_material(db, type_, spec, size, uom):

    type_ = normalize(type_)
    spec = normalize(spec)
    size = normalize(size)

    type_key = (type_ or "").lower()
    spec_key = (spec or "").lower()

    name = ", ".join(
        x for x in (type_, spec) if x
    ).upper()

    rm = db.scalar(
        select(RawMaterial).where(
            func.lower(func.coalesce(RawMaterial.type, "")) == type_key,
            func.lower(func.coalesce(RawMaterial.spec, "")) == spec_key,
        )
    )

    if rm:
        rm.size_text = size or rm.size_text
        rm.uom = uom or rm.uom
        return rm

    rm = RawMaterial(
        code=next_code(db, RawMaterial, "code", prefix="M", width=4),
        name=name,
        type=type_,
        spec=spec,
        size_text=size,
        uom=uom,
    )

    db.add(rm)
    db.flush()
    return rm


def upsert_material_po_line(db, mpo, part_no, rm, qty, size, length):
    line = db.scalar(
        select(MaterialPOLine).where(
            MaterialPOLine.po_id == mpo.id,
            MaterialPOLine.part_no == part_no,
            MaterialPOLine.material_id == rm.id,
        )
    )
    # print("\t", part_no, qty, size, length)
    if line:
        # ✅ UPDATE existing line
        if qty is not None:
            line.qty_ordered = qty
        if size:
            line.size = size
        if length:
            line.length_text = length
        return line

    # ✅ INSERT new line
    line = MaterialPOLine(
        po_id=mpo.id,
        part_no=part_no,
        material_id=rm.id,
        qty_ordered=qty if qty is not None else Decimal(0),
        size=size,
        length_text=length,
    )
    db.add(line)
    db.flush()
    return line


# =====================================================
# ============== RAW BATCH (RECEIVING) ================
# =====================================================

def upsert_raw_batch(
    db,
    batch_no,
    rm,
    mpo,
    line,
    heat_no,
    size_text,
    length_text,
    weight,
    supplier_id=None,
    cutting_note=None,
    po_note=None,
):
    rb = db.scalar(
        select(RawBatch).where(
            RawBatch.material_id == rm.id,
            RawBatch.batch_no == batch_no,
        )
    )

    if rb:
        # ✅ Update references (safe)
        rb.po_id = rb.po_id or mpo.id
        rb.material_po_line_id = line.id
        # 👇 เพิ่ม
        rb.supplier_id = supplier_id

        # ✅ Update metadata
        if heat_no:
            rb.mill_heat_no = heat_no
        if length_text:
            rb.length_text = length_text

        # ✅ Accumulate weight ONLY if provided
        if weight is not None:
            rb.qty_received = (rb.qty_received or Decimal(0)) + weight
            rb.weight = (rb.weight or Decimal(0)) + weight

        db.flush()
        return rb

    # ✅ INSERT new batch
    rb = RawBatch(
        material_id=rm.id,
        supplier_id=supplier_id,

        po_id=mpo.id,
        material_po_line_id=line.id,

        batch_no=batch_no,

        mill_heat_no=heat_no,
        heat_lot=heat_no,

        size_text=size_text,
        length_text=length_text,

        cutting_note=cutting_note,
        po_note=po_note,

        weight=weight or Decimal(0),
        qty_received=weight or Decimal(0),
    )
    db.add(rb)
    db.flush()
    return rb


def upsert_part_material(db, part_id, raw_material_id):
    pm = db.scalar(
        select(PartMaterial).where(
            PartMaterial.part_id == part_id,
            PartMaterial.raw_material_id == raw_material_id,
        )
    )

    if not pm:
        pm = PartMaterial(
            part_id=part_id,
            raw_material_id=raw_material_id,
        )
        db.add(pm)

    # future fields go here
    # pm.is_active = True

    db.flush()
    return pm

# =====================================================
# ============ LOT MATERIAL USE =======================
# =====================================================

def get_lot_by_po(db, po_no):
    return db.scalar(
        select(ProductionLot)
        .join(ProductionLot.po)
        .where(ProductionLot.po.has(po_number=po_no))
        .order_by(ProductionLot.created_at.desc())
    )


def get_remaining_batch_qty(db, rb):
    used = db.scalar(
        select(func.coalesce(func.sum(LotMaterialUse.qty), 0))
        .where(LotMaterialUse.batch_id == rb.id)
    )
    return (rb.qty_received or Decimal(0)) - used


def insert_lot_material_use(db, lot, rb, qty, note=None):
   
    if not qty or qty <= 0:
        return None

    remaining = get_remaining_batch_qty(db, rb)
    if remaining <= 0:
        return None

    use_qty = min(Decimal(qty), Decimal(remaining))

    # ✅ Prevent duplicate import (same lot + batch + qty + note)
    existing = db.scalar(
        select(LotMaterialUse)
        .where(
            LotMaterialUse.lot_id == lot.id,
            LotMaterialUse.batch_id == rb.id,
            LotMaterialUse.qty == use_qty,
            LotMaterialUse.note == note,
        )
    )
    if existing:
        return existing

    lmu = LotMaterialUse(
        lot_id=lot.id,
        batch_id=rb.id,
        raw_material_id=rb.material_id,
        qty=use_qty,
        qty_uom="pcs",
        note=note,
    )
    db.add(lmu)
    db.flush()
    return lmu

def deleteData(db):
    from sqlalchemy import delete

    db.execute(delete(LotMaterialUseHistory))

    db.execute(delete(LotMaterialUse))

    # ลบ Reference ก่อน
    db.execute(delete(RawBatchReference))

    db.execute(delete(RawBatch))
    db.execute(delete(MaterialPOLine))
    db.execute(delete(MaterialPO))
    db.execute(delete(PartMaterial))
    db.execute(delete(RawMaterial))

    db.commit()


from models import RawBatchReference, PO

def upsert_raw_batch_reference(
    db,
    batch,
    part,
    po_no,
    lot=None,
):
    po = None

    if po_no:
        po = db.scalar(
            select(PO).where(
                PO.po_number == po_no
            )
        )

    rev_id = None

    if lot:
        rev_id = lot.part_revision_id

    ref = db.scalar(
        select(RawBatchReference).where(
            RawBatchReference.batch_id == batch.id,
            RawBatchReference.part_id == part.id,
            RawBatchReference.part_revision_id == rev_id,
            RawBatchReference.po_id == (po.id if po else None),
            RawBatchReference.lot_id == (lot.id if lot else None),
        )
    )

    if ref:
        return ref

    ref = RawBatchReference(
        batch_id=batch.id,
        part_id=part.id,
        part_revision_id=rev_id,
        po_id=po.id if po else None,
        lot_id=lot.id if lot else None,
    )

    db.add(ref)
    db.flush()

    return ref
# =====================================================
# =================== MAIN ============================
# =====================================================

XLSX_FILE = Path(
    # r"Z:\Topnotch Group\Public\2026\Material Cert 2026\Material 2026.xls"
    
    # r"Z:\Topnotch Group\Public\Testing APP\importMat\ALL MATERIAL.xlsx"
    r"Z:\Topnotch Group\Public\Testing APP\importMat\Material 2026.xlsx"
)
def main():

    from openpyxl import load_workbook

    with SessionLocal() as db:
        # deleteData(db)
        wb = load_workbook(
            XLSX_FILE,
            data_only=True,
            read_only=True
        )

        # ws = wb.active
        ws = wb["Order"]
        # ws = wb["Sort by size"]
        # 

        headers = [
            str(c.value).strip() if c.value else ""
            for c in ws[1]
        ]

        count = 0

        for values in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            row = dict(zip(headers, values))

            if not any(values):
                continue

            company = normalize(
                row.get("Company")
            )

            vendor_po = (
                normalize(row.get("Vendor PO"))
                or normalize(row.get("Heat lot"))
            )

            part_no = normalize(
                row.get("Part no.")
            )

            size = row.get("Size")
            lenght = row.get("Length")

            if not company:

                if vendor_po:
                    company = f"UNKNOWN-{vendor_po}"
                else:
                    company = "UNKNOWN-NO-PO"

            supplier = upsert_supplier(
                db,
                company
            )

            rm = upsert_raw_material(
                db,
                normalize(row.get("Type")),
                normalize(row.get("Spec.") or row.get("Spec")),
                normalize(row.get("Size")),
                "feet"
            )

            mpo = upsert_material_po(
                db,
                supplier,
                vendor_po,
                parse_date(row.get("Date"))
            )

            po_items = parse_po_qty(
                row.get("PO#, Qty")
            )

            for it in po_items:

                if not part_no:
                    print(
                        "⚠️ SKIP row: missing Part no."
                    )
                    continue

                part_items = part_no.split()

                for pp_no in part_items:

                    part = db.scalar(
                        select(Part).where(
                            Part.part_no == pp_no
                        )
                    )

                    if not part:
                        continue

                    upsert_part_material(
                        db,
                        part.id,
                        rm.id
                    )

                    line = upsert_material_po_line(
                        db,
                        mpo,
                        part_no,
                        rm,
                        it["qty"],
                        size,
                        lenght,
                    )

                    rb = upsert_raw_batch(
                        db=db,
                        batch_no=vendor_po,
                        rm=rm,
                        mpo=mpo,
                        line=line,

                        heat_no=normalize(row.get("Heat lot")),
                        size_text=normalize(row.get("Size")),
                        length_text=normalize(row.get("Length")),
                        weight=parse_decimal(it["qty"]),

                        supplier_id=supplier.id,

                        cutting_note=row.get("Cutting Receiving/HT"),
                        po_note=row.get("PO#, Qty"),
                    )

                    lot = get_lot_by_po(
                        db,
                        it["po"]
                    )

                    upsert_raw_batch_reference(
                        db=db,
                        batch=rb,
                        part=part,
                        po_no=it["po"],
                        lot=lot,
                    )

                    if not lot:
                        continue

                    insert_lot_material_use(
                        db,
                        lot=lot,
                        rb=rb,
                        qty=it["qty"],
                        note=it["note"],
                    )

            count += 1

            if count % 50 == 0:
                db.commit()
                print(
                    f"✅ {count} rows processed"
                )

        db.commit()

        print(
            f"🎉 DONE: {count} rows imported"
        )


if __name__ == "__main__":
    main()
